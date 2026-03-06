from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import date
import re

from configuracion.models import Colegio, Profesor, Libro
from .models import Bloque, Clase, Asignacion, ClaseParticular

# ─────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────
DIAS_ES  = {0:'Lun', 1:'Mar', 2:'Mié', 3:'Jue', 4:'Vie', 5:'Sáb', 6:'Dom'}
MESES_ES = {1:'Ene', 2:'Feb', 3:'Mar', 4:'Abr', 5:'May', 6:'Jun',
            7:'Jul', 8:'Ago', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dic'}
UNIDADES_ESPECIALES = {'A': 'Material Asignado', 'S': 'Socialización de Simulacro'}

# ─────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────

def _fecha_label(d):
    return f"{DIAS_ES[d.weekday()]} {d.day:02d}-{MESES_ES[d.month]}"

def _extraer_num_grado(g):
    nums = re.findall(r'\d+', g)
    return int(nums[0]) if nums else 0

def _fill(color):
    return PatternFill('solid', fgColor=color)

def _font(bold=False, link=False, size=10):
    kwargs = {'name': 'Arial', 'size': size, 'bold': bold}
    if link:
        kwargs['color'] = '0563C1'
        kwargs['underline'] = 'single'
    return Font(**kwargs)


def _build_asignaciones_map(colegio_ids):
    """
    Pre-carga TODAS las asignaciones de los colegios indicados y las agrupa
    en un dict: (colegio_id, grado) → [Asignacion, ...]

    Esto evita el problema N+1 de hacer una consulta por cada clase.
    """
    resultado = defaultdict(list)
    for a in Asignacion.objects.filter(colegio_id__in=colegio_ids):
        resultado[(a.colegio_id, a.grado)].append(a)
    return resultado


def _titulo_libro(asig_map, colegio_id, grado, fecha):
    """
    Busca en el map pre-cargado la asignación cuya ventana de fechas
    cubre la fecha dada. Devuelve '' si ninguna la cubre.
    """
    for a in asig_map.get((colegio_id, grado), []):
        inicio = a.fecha_inicio
        fin    = a.fecha_fin
        if inicio and fin and inicio <= fecha <= fin:
            return a.libro_titulo
        if inicio is None and fin is None:
            return a.libro_titulo
    return ''


# ─────────────────────────────────────────────────────────────
# DESCARGA EXCEL — COLEGIO
# ─────────────────────────────────────────────────────────────

def descargar_excel_colegio(request, colegio_id):
    colegio = get_object_or_404(Colegio, id=colegio_id)
    año = date.today().year

    clases = list(
        Clase.objects
        .filter(colegio=colegio, fecha__year=año)
        .select_related('bloque', 'profesor')
        .order_by('fecha', 'bloque__orden')
    )
    fechas = sorted(set(c.fecha for c in clases))

    # Matriz bloque_id → fecha → clase
    matriz = defaultdict(dict)
    for c in clases:
        matriz[c.bloque_id][c.fecha] = c

    # Bloques agrupados por grado
    bloques = list(Bloque.objects.filter(colegio=colegio).order_by('grado', 'orden'))
    grados_dict = defaultdict(list)
    for b in bloques:
        grados_dict[b.grado].append(b)
    grados_ordenados = sorted(grados_dict.keys(), key=_extraer_num_grado, reverse=True)

    # ── CORRECCIÓN 1: clave incluye libro_titulo para evitar colisiones
    # entre libros distintos que comparten materia y número de unidad.
    libros_map = {
        (l.titulo, l.materia, str(l.unidad)): l
        for l in Libro.objects.all()
    }

    # ── CORRECCIÓN 2: pre-cargar asignaciones en memoria; sin N+1.
    asig_map = _build_asignaciones_map([colegio.id])

    # ── Construir workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = colegio.nombre[:31]

    ws['A1'] = '🏫'; ws['A1'].font = _font(bold=True)
    ws['B1'] = '⏱️'; ws['B1'].font = _font()
    ws['C1'] = 'GRADO - SALÓN'; ws['C1'].font = _font(bold=True)

    for ci, f in enumerate(fechas, start=4):
        cell = ws.cell(2, ci, _fecha_label(f))
        cell.font = _font(bold=True)
        cell.fill = _fill('D0D0D0')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    fila = 3
    for grado in grados_ordenados:
        for bloque in sorted(grados_dict[grado], key=lambda x: x.orden):

            for ci_fix, val in [(1, colegio.nombre), (2, bloque.hora), (3, grado)]:
                c = ws.cell(fila, ci_fix, val)
                c.font = _font()
                c.fill = _fill('D4D4D4')

            for ci, fecha in enumerate(fechas, start=4):
                clase = matriz[bloque.id].get(fecha)
                if not clase:
                    continue

                if clase.cancelada:
                    ws.cell(fila, ci, 'CANCELADA').font = _font()
                elif clase.es_evento:
                    ws.cell(fila, ci, clase.titulo_evento or 'Evento').font = _font()
                else:
                    nombre_profe = clase.profesor.nombre_corto if clase.profesor else ''
                    ws.cell(fila, ci, nombre_profe).font = _font()

                if clase.cancelada or clase.es_evento or not clase.materia:
                    continue

                unidad = str(clase.unidad) if clase.unidad else ''

                if unidad in UNIDADES_ESPECIALES:
                    texto = f'{clase.materia} | {UNIDADES_ESPECIALES[unidad]}'
                    ws.cell(fila + 1, ci, texto).font = _font()

                elif unidad.isdigit():
                    # ── CORRECCIÓN: lookup con (titulo, materia, unidad)
                    titulo = _titulo_libro(asig_map, colegio.id, grado, fecha)
                    libro_obj = libros_map.get((titulo, clase.materia, unidad))
                    nom_u = libro_obj.nombre_unidad if libro_obj else ''
                    unidad_txt = f'U.{unidad}: {nom_u}' if nom_u else f'U.{unidad}'
                    texto = f'{titulo} | {clase.materia} | {unidad_txt}' if titulo else f'{clase.materia} | {unidad_txt}'
                    ws.cell(fila + 1, ci, texto).font = _font()

                else:
                    titulo = _titulo_libro(asig_map, colegio.id, grado, fecha)
                    texto = f'{titulo} | {clase.materia}' if titulo else clase.materia
                    ws.cell(fila + 1, ci, texto).font = _font()

            fila += 2

    fila += 1

    # ── Tabla de docentes por materia ──
    # Agrupamos: (profesor_id, materia) → {grado: count}
    # Así cada fila representa UN profesor + UNA materia, no el total.
    conteo = defaultdict(lambda: defaultdict(int))   # (prof_id, materia) → {grado: n}
    ids_profes = set()
    for c in clases:
        if c.profesor and not c.cancelada and not c.es_evento and c.materia:
            conteo[(c.profesor_id, c.materia)][c.bloque.grado] += 1
            ids_profes.add(c.profesor_id)

    # Cabecera
    for ci_h, label in [(1, 'DOCENTES'), (2, 'DOCUMENTO'), (3, 'MATERIA')]:
        cell = ws.cell(fila, ci_h, label)
        cell.font = _font(bold=True)
        cell.fill = _fill('FCE4D6')

    for ci_g, grado in enumerate(grados_ordenados, start=4):
        cell = ws.cell(fila, ci_g, f'Clases {grado}')
        cell.font = _font(bold=True)
        cell.fill = _fill('FCE4D6')

    fila += 1

    # Ordenamos por nombre de profesor y luego por materia
    profes_map = {p.id: p for p in Profesor.objects.filter(id__in=ids_profes)}
    filas_datos = sorted(
        conteo.keys(),
        key=lambda k: (profes_map[k[0]].nombre_corto, k[1])
    )

    profe_anterior = None
    for (prof_id, materia) in filas_datos:
        profe = profes_map[prof_id]
        # Solo imprime nombre y documento en la primera fila del profesor
        if prof_id != profe_anterior:
            ws.cell(fila, 1, profe.nombre_corto).font = _font(bold=True)
            ws.cell(fila, 2, profe.documento or '').font = _font()
            profe_anterior = prof_id
        else:
            ws.cell(fila, 1, '').font = _font()
            ws.cell(fila, 2, '').font = _font()

        ws.cell(fila, 3, materia).font = _font()

        for ci_g, grado in enumerate(grados_ordenados, start=4):
            cnt = conteo[(prof_id, materia)].get(grado, 0)
            ws.cell(fila, ci_g, cnt or '').font = _font()

        fila += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    for ci in range(4, 4 + len(fechas)):
        ws.column_dimensions[get_column_letter(ci)].width = 38

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre = f"Colegio_{colegio.nombre.replace(' ', '_')}_{año}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────
# DESCARGA EXCEL — PROFESOR
# Incluye Clases de colegios + ClaseParticular, ordenadas por fecha y hora
# ─────────────────────────────────────────────────────────────

def descargar_excel_profesor(request, profesor_id):
    profesor = get_object_or_404(Profesor, id=profesor_id)
    año = date.today().year

    def _minutos(hora_str):
        try:
            partes = hora_str.split('-')[0].strip().split(':')
            return int(partes[0]) * 60 + int(partes[1])
        except Exception:
            return 9999

    # ── 1. Normalizar ambos tipos de clase en una lista unificada ──
    # Cada entrada es un dict con los mismos campos para renderizar igual.

    entradas = []

    clases_colegio = list(
        Clase.objects
        .filter(profesor=profesor, fecha__year=año, cancelada=False, es_evento=False)
        .select_related('bloque__colegio', 'bloque')
    )

    # Pre-cargar asignaciones y libros en memoria
    colegios_ids = list({c.bloque.colegio_id for c in clases_colegio})
    asig_map   = _build_asignaciones_map(colegios_ids)
    libros_map = {(l.titulo, l.materia, str(l.unidad)): l for l in Libro.objects.all()}

    for c in clases_colegio:
        colegio = c.bloque.colegio
        grado   = c.bloque.grado
        titulo  = _titulo_libro(asig_map, colegio.id, grado, c.fecha)
        unidad_c = str(c.unidad) if c.unidad else ''
        material_c = UNIDADES_ESPECIALES.get(unidad_c, titulo or 'Sin libro asignado')
        entradas.append({
            'fecha':          c.fecha,
            'hora':           c.bloque.hora,
            'minutos':        _minutos(c.bloque.hora),
            'colegio_nombre': colegio.nombre,
            'ciudad':         colegio.ciudad or '',
            'mapa_link':      colegio.mapa_link or '',
            'grado':          grado,
            'material':       material_c,
            'materia':        c.materia or '',
            'unidad':         str(c.unidad) if c.unidad else '',
            'tipo':           'colegio',
            'libro_obj':      libros_map.get((titulo, c.materia, str(c.unidad))) if c.unidad else None,
        })

    particulares = list(
        ClaseParticular.objects
        .filter(profesor=profesor, fecha__year=año)
    )

    for p in particulares:
        unidad = str(p.unidad) if p.unidad else ''
        libro_obj = libros_map.get((p.material, p.materia, unidad))
        material_p = UNIDADES_ESPECIALES.get(unidad, p.material or 'Sin libro asignado')
        entradas.append({
            'fecha':          p.fecha,
            'hora':           p.hora,
            'minutos':        _minutos(p.hora),
            'colegio_nombre': p.estudiante,   # estudiante/colegio particular
            'ciudad':         p.ciudad or '',
            'mapa_link':      p.mapa_link or '',
            'grado':          p.grado,
            'material':       material_p,
            'materia':        p.materia,
            'unidad':         unidad,
            'tipo':           'particular',
            'libro_obj':      libro_obj,
        })

    # Ordenar por fecha y luego por hora
    entradas.sort(key=lambda e: (e['fecha'], e['minutos']))

    wb = Workbook()
    ws = wb.active
    ws.title = profesor.nombre_corto[:31]

    if not entradas:
        ws['A1'] = f'{profesor.nombre} — Sin clases registradas en {año}'
        ws['A1'].font = _font(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Profesor_{profesor.nombre_corto}.xlsx"'
        wb.save(response)
        return response

    # ── 2. Fila 1: nombre del profesor + fechas (combinar celdas del mismo día) ──
    ws['A1'] = profesor.nombre
    ws['A1'].font = _font(bold=True)
    ws['A1'].fill = _fill('B0B3B2')

    for ci, entrada in enumerate(entradas, start=2):
        cell = ws.cell(1, ci, _fecha_label(entrada['fecha']))
        cell.font = _font(bold=True)
        cell.fill = _fill('D9D9D9')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Combinar celdas consecutivas con la misma fecha
    grupos_fecha = []
    ci_inicio = 2
    for i, entrada in enumerate(entradas):
        ci_actual = i + 2
        fecha_sig = entradas[i + 1]['fecha'] if i < len(entradas) - 1 else None
        if fecha_sig != entrada['fecha']:
            grupos_fecha.append((ci_inicio, ci_actual, entrada['fecha']))
            ci_inicio = ci_actual + 1

    for (col_ini, col_fin, fecha) in grupos_fecha:
        if col_ini < col_fin:
            ws.merge_cells(start_row=1, start_column=col_ini,
                           end_row=1,   end_column=col_fin)
            cell = ws.cell(1, col_ini, _fecha_label(fecha))
            cell.font = _font(bold=True)
            cell.fill = _fill('D9D9D9')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── 3. Filas 2-8: etiquetas en columna A ──
    labels = ['Colegio', 'Ciudad', 'Hora', 'Grado', 'Material', 'Asignatura', 'Unidad']
    for row_i, label in enumerate(labels, start=2):
        cell = ws.cell(row_i, 1, label)
        cell.font = _font(bold=True)
        cell.fill = _fill('D4D4D4')

    # ── 4. Columnas de datos (una por entrada) ──
    for ci, e in enumerate(entradas, start=2):
        # Fila 2: colegio/estudiante con hyperlink de Maps si existe
        cell_col = ws.cell(2, ci, e['colegio_nombre'])
        if e['mapa_link']:
            cell_col.hyperlink = e['mapa_link']
            cell_col.font = _font(link=True)
        else:
            cell_col.font = _font()

        # Clases particulares: fondo amarillo suave para distinguirlas
        fill_particular = _fill('FFF9C4') if e['tipo'] == 'particular' else None
        if fill_particular:
            for row_i in range(2, 9):
                ws.cell(row_i, ci).fill = fill_particular

        ws.cell(3, ci, e['ciudad']).font  = _font()
        ws.cell(4, ci, e['hora']).font    = _font()
        ws.cell(5, ci, e['grado']).font   = _font()
        ws.cell(6, ci, e['material']).font = _font()
        ws.cell(7, ci, e['materia']).font  = _font()

        # Fila 8: unidad con hyperlink al link del libro si existe
        unidad = e['unidad']
        if unidad in UNIDADES_ESPECIALES:
            ws.cell(8, ci, UNIDADES_ESPECIALES[unidad]).font = _font()
        elif unidad.isdigit() and e['libro_obj']:
            libro_obj = e['libro_obj']
            nom_u = libro_obj.nombre_unidad or ''
            unidad_txt = f'{unidad}. {nom_u}' if nom_u else unidad
            cell_u = ws.cell(8, ci, unidad_txt)
            if libro_obj.link_unidad:
                cell_u.hyperlink = libro_obj.link_unidad
                cell_u.font = _font(link=True)
            else:
                cell_u.font = _font()
        else:
            ws.cell(8, ci, unidad).font = _font()

    ws.column_dimensions['A'].width = 12
    for ci in range(2, 2 + len(entradas)):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    ws.row_dimensions[1].height = 30

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre = f"Profesor_{profesor.nombre_corto.replace(' ', '_')}_{año}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response
